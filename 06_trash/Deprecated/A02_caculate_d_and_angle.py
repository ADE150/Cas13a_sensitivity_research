import os.path

from common_tools import *
import openpyxl

"""
本脚本的任务：

1. 把WT2_1中的963和967附近的残基（不包括碱基）选择出来，保存到新文件「wt2_1_963_967_around_5」中
2. 把每帧文件的距离和二面角计算出来，导出到文件中，由下一个脚本绘制表格

从本文件开始，输入和输出路径统一在常量文件里，使用的基础方法统一在z_tool文件里，每个脚本不用写成方法，只需用命令书写方式完成任务即可
"""


# 所有原始PDB文件的文件夹路径
input_origin_path = origin_PDB_path
# 所有原始PDB文件的文件名称（有后缀）
input_origin_filename_list = [f for f in os.listdir(input_origin_path) if f.endswith(".pdb")]
input_origin_filename_list.sort(key=lambda x: (x[:3], int(find_first_str(x, r"_(.+?)\.pdb"))))
# 所有原始PDB文件的文件路径
input_origin_dir_list = [con_frag(input_origin_path, f) for f in input_origin_filename_list]
# 所有原始对象的名称
all_origin_object_name=[get_filename_without_suffix(f) for f in input_origin_filename_list]
# 对象集合文件的文件路径
output_all_object_dir = con_frag(intermediate_products_dir, distance_and_angle_folder + r"\all_object.pdb")
# 「选择附近残基」文件的文件路径
output_resis_around_object_dir = con_frag(intermediate_products_dir,
                                          distance_and_angle_folder + r"\resis_around.pdb")
# 距离和二面角结果的文件路径
output_distance_and_angle_result_dir = con_frag(output_dir,
                                                distance_and_angle_folder + r"\distance_and_angle.xlsx")
# # 启动Pymol
# pymol.finish_launching()
#
# # 加载所有PDB集中到一个文件中并保存
# remind_progress("加载所有PDB到一个文件中")
# for path in input_origin_paths:
#     the_object_name = get_filename_without_suffix(path)
#     pymol_cmd.load(path, the_object_name)
#     remove_solvent(the_object_name)
#     pymol_cmd.disable(the_object_name)
#     remind_detail(the_object_name)


# # 生成WT2_1的选择残基文件
# pymol_cmd.reinitialize()
# input_resis_around_dir = connect_dir(input_origin_path , r"wt2_1.pdb")
# wt2_1_object_name = "wt2_1"
# pymol_cmd.load(input_resis_around_dir, wt2_1_object_name)
# remove_solvent()
# wt2_1_963 = "wt2_1_R963"
# wt2_1_967 = "wt2_1_Y967"
# wt2_1_963_around_5 = "wt2_1_R963_around_5"
# wt2_1_967_around_5 = "wt2_1_Y967_around_5"
# pymol_cmd.select(wt2_1_963, f"resi 963 and chain A")
# pymol_cmd.select(wt2_1_967, f" resi 967 and chain A")
# pymol_cmd.select(wt2_1_963_around_5, f"byres resi 963 around 5 and chain A")
# pymol_cmd.select(wt2_1_967_around_5, f"byres resi 967 around 5 and chain A")
# pymol_cmd_save(output_resis_around_object_dir)
# pymol_cmd.quit()


# # 计算残基-质心距离和A(-3)和A(-4)的二面角
# # remind_progress("重新加载对象集合")
# # 创建用于保存距离和二面角的工作簿
# distance_and_angle_workbook=openpyxl.Workbook()
# worksheet_1=distance_and_angle_workbook.active
# worksheet_1.title = 'Distances and Angles'
# # 写入表头
# worksheet_1.append(['对象', '距离 (Å)', '二面角 (°)'])
#
# # pymol_cmd.load(output_all_object_dir)
# for selection in origin_object_names:
#     # 选择链A的963位Cα原子
#     ca_A963 = f'object {selection} and chain A and resi 963 and name CA'
#     # 选择链B的28位A碱基的N1和C6原子
#     atoms_B28 = f'object {selection} and chain B and resi 28 and name N1+C6'
#     # 创建一个伪原子来代表B链28位A碱基N1-C6环的质心
#     object_name_com_28=f'{selection}_com_B28'
#     pymol_cmd.pseudoatom(object_name_com_28, selection=atoms_B28)
#     # 计算A链963位Cα原子与B链28位A碱基质心之间的距离
#     distance = pymol_cmd.get_distance(ca_A963, object_name_com_28)
#     # 选择链B 27位A碱基的C3'原子
#     atom_B27_C3 = f'object {selection} and chain B and resi 27 and name C3\''
#     # 选择链B 28位A碱基的C3'原子
#     atom_B28_C3 = f'object {selection} and chain B and resi 28 and name C3\''
#     # 选择链B 27位A碱基的C8原子
#     atom_B27_C8 = f'object {selection} and chain B and resi 27 and name C8'
#     # 选择链B 27位A碱基的C2原子
#     atom_B27_C2 = f'object {selection} and chain B and resi 27 and name C2'
#     # 计算二面角 θ
#     dihedral_angle = pymol_cmd.get_dihedral(atom_B27_C3, atom_B28_C3, atom_B27_C8, atom_B27_C2)
#     # 写入Excel工作表
#     worksheet_1.append([selection, distance, dihedral_angle])
#
# distance_and_angle_workbook.save(output_distance_and_angle_result_dir)
#
# pymol_cmd_save(output_all_object_dir)
#
#
# # 关闭Pymol
# pymol_cmd.quit()

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# 读取Excel文件
# filename = 'distances_and_angles.xlsx'
filename = output_distance_and_angle_result_dir
workbook = openpyxl.load_workbook(filename)
worksheet = workbook.active

# 提取数据
objects = []
distances = []
angles = []

# for row in worksheet.iter_rows(min_row=2, values_only=True):
#     objects.append(row[0])
#     distances.append(row[1])
#     angles.append(row[2])
#
# # 将数据分为WT2和MT3两组
# wt2_distances = []
# mt3_distances = []
# wt2_angles = []
# mt3_angles = []
#
# for n, dist, angle in zip(objects, distances, angles):
#     if n.startswith('wt2'):
#         wt2_distances.append(dist)
#         wt2_angles.append(angle)
#     elif n.startswith('mt3'):
#         mt3_distances.append(dist)
#         mt3_angles.append(angle)
#
# # 创建极坐标图
# plt.figure(figsize=(14, 7))
#
# ax1 = plt.subplot(121, projection='polar')
#
# # 将角度转换为弧度
# wt2_angles_rad = np.deg2rad(wt2_angles)
# mt3_angles_rad = np.deg2rad(mt3_angles)
#
# # 绘制极坐标图
# ax1.scatter(wt2_angles_rad, wt2_distances, color='#bebebe', label='WT2')
# ax1.scatter(mt3_angles_rad, mt3_distances, color='#3fe0d0', label='MT3')
#
# # 设置极坐标网格
# ax1.set_rmax(6)
# ax1.set_rticks([2, 4, 6])
# ax1.set_rlabel_position(-22.5)  # Move radial labels away from plotted line
# ax1.grid(True)
# ax1.set_title("Polar Plot of Distance and Dihedral Angle")
#
# # 创建时间演化图
# times = np.arange(0, 100, 5)  # 每帧间隔5ps，共20帧，跨度为100ps
#
# ax2 = plt.subplot(222)
# ax3 = plt.subplot(224)
#
# # 绘制距离随时间变化的图
# ax2.plot(times, wt2_distances, color='#bebebe', label='WT2')
# ax2.plot(times, mt3_distances, color='#3fe0d0', label='MT3')
# ax2.set_xlabel("Time / ps")
# ax2.set_ylabel("Distance d / Å")
# ax2.set_ylim(2, 6)
# ax2.set_title("Distance Evolution Over Time")
# ax2.legend()
#
# # 绘制角度随时间变化的图
# ax3.plot(times, wt2_angles, color='#bebebe', label='WT2')
# ax3.plot(times, mt3_angles, color='#3fe0d0', label='MT3')
# ax3.set_xlabel("Time / ps")
# ax3.set_ylabel("Angle θ / °")
# ax3.set_ylim(-120, -60)
# ax3.set_title("Angle Evolution Over Time")
# ax3.legend()
#
# # 调整布局
# plt.tight_layout()
# plt.savefig('distance_angle_evolution.png')
# plt.show()

# ————  跑图代码02  ————

# 读取Excel文件
# filename = 'distances_and_angles.xlsx'
# workbook = openpyxl.load_workbook(filename)
# worksheet = workbook.active
#
# # 提取数据
# objects = []
# distances = []
# angles = []

for row in worksheet.iter_rows(min_row=2, values_only=True):
    objects.append(row[0])
    distances.append(row[1])
    angles.append(row[2])

# 将数据分为WT2和MT3两组
wt2_distances = []
mt3_distances = []
wt2_angles = []
mt3_angles = []

for obj, dist, angle in zip(objects, distances, angles):
    if obj.startswith('wt2'):
        wt2_distances.append(dist)
        wt2_angles.append(angle)
    elif obj.startswith('mt3'):
        mt3_distances.append(dist)
        mt3_angles.append(angle)

# 创建极坐标图
plt.figure(figsize=(14, 7))

ax1 = plt.subplot(121, projection='polar')

# 设置极坐标图网格线的颜色
ax1.grid(color='#f0f0f0',linewidth=1.75)
# 设置极坐标图外围圆的颜色
ax1.spines['polar'].set_edgecolor('#f0f0f0')
# 将角度转换为弧度
wt2_angles_rad = np.deg2rad(wt2_angles)
mt3_angles_rad = np.deg2rad(mt3_angles)

# 绘制极坐标图
ax1.scatter(wt2_angles_rad, wt2_distances, color='#bebebe', label='WT2')
ax1.scatter(mt3_angles_rad, mt3_distances, color='#3fe0d0', label='MT3')

# 设置极坐标网格
ax1.set_rmax(6)
# ax1.set_rticks([0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5, 5.5, 6])
ax1.set_rticks([2,4,6])
ax1.set_theta_zero_location('E')  # 0°在右边
ax1.set_thetamin(-180)  # 设置极坐标的最小角度
ax1.set_thetamax(180)  # 设置极坐标的最大角度
ax1.set_thetagrids(range(-180, 180, 45))  # 显示角度网格

ax1.grid(True)
ax1.set_title("Polar Plot of Distance and Dihedral Angle")

# 创建时间演化图
times = np.arange(0, 100, 5)  # 每帧间隔5ps，共20帧，跨度为100ps

ax2 = plt.subplot(222)
ax3 = plt.subplot(224)

# 绘制距离随时间变化的图
ax2.plot(times, wt2_distances, color='#bebebe', label='WT2')
ax2.plot(times, mt3_distances, color='#3fe0d0', label='MT3')
ax2.set_xlabel("Time / ps")
ax2.set_ylabel("Distance d / Å")
ax2.set_ylim(2, 6)
ax2.set_title("Distance Evolution Over Time")
ax2.legend()

# 绘制角度随时间变化的图
ax3.plot(times, wt2_angles, color='#bebebe', label='WT2')
ax3.plot(times, mt3_angles, color='#3fe0d0', label='MT3')
ax3.set_xlabel("Time / ps")
ax3.set_ylabel("Angle θ / °")
ax3.set_ylim(-120, -60)
ax3.set_title("Angle Evolution Over Time")
ax3.legend()

# 调整布局和边距
plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1, wspace=0.4, hspace=0.4)

# 调整布局
# plt.tight_layout()
output_distance_and_angle_result_plot_dir=con_frag(os.path.dirname(output_distance_and_angle_result_dir), r"distance_angle_plot.png")
print_stack_trace(output_distance_and_angle_result_plot_dir)
plt.savefig(output_distance_and_angle_result_plot_dir)
plt.show()





